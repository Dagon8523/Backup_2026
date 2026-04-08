# -*- coding: utf-8 -*-
"""
Matriz de Riesgos -> Excel limpio para SAP Datasphere
- Encabezados: fila 2 (1-index).
- Forward-fill en TODAS las columnas (resolver celdas combinadas) **excepto** SUBSISTEMA_A_LAS_QUE_PERTENECE_EL_RIESGO.
- Limpieza de tildes/acentos, espacios, símbolos y no-ASCII.
- Filtra filas: mantener SOLO las que tengan 'Descripcion del Control' no vacía.
- DEDUPE: por cada 'No' dejar UNA fila por cada descripcion distinta.
- %: Detecta columnas de porcentaje y normaliza a número 0..100 con 2 decimales (sin %), conservando '.' como separador decimal.
- Renombrado final de columnas en el orden indicado por el usuario.
- Reescribe con xlsxwriter (sin imágenes/objetos).
"""

from pathlib import Path
import re
from typing import List, Optional
import pandas as pd
from openpyxl import load_workbook
from unidecode import unidecode

# ============================ Config ============================
BASE = Path(r"C:\Users\degonzalez\Documents\2025\DataSphere\Calidad")
INPUT_XLSX  = BASE / r"Matriz_de_Riesgos_CONSOLIDADA_FINAL.xlsx"
SHEET_NAME  = "Matriz de Riesgos"
HEADER_ROW_NUMBER = 2
OUTPUT_XLSX = BASE / "matriz_riesgos_procesada.xlsx"

# Opciones de limpieza
PRESERVAR_ENYE = False   # False => 'ñ'->'n'
COLLAPSE_SPACES = True

# Modo de renombrado final:
RENAME_STRICT = True  # True => error si el número de columnas != columnas_finales

# === Lista final de nombres (izquierda -> derecha) =================
COLUMNAS_FINALES = [
    "NO","PROCESO","SUBPROCESO","AREA_DE_TRABAJO","OBJETIVO_ESTRATEGICO","RIESGO_ESTRATEGICO",
    "EVENTO_DE_RIESGO_ME_PODRIA_GENERAR","VULNERABILIDAD_DEBILIDAD_ASOCIADA","AMENAZAS_ACCION_O_FENOMENO",
    "CAUSAS","CONSECUENCIAS_POTENCIALES","SUBSISTEMA_A_LAS_QUE_PERTENECE_EL_RIESGO",
    "TIPO_DE_RIESGO_VER_TIPOS_DE_RIEGOS","CATEGORIA_DE_RIESGO_VER_CATEGORIAS_DE_RIESGO",
    "FUENTES_DE_RIESGO_VER_FUENTES_DE_RIESGO","TIPOS_DE_IMPACTO_VER_TIPOS_DE_IMPACTO",
    "FRECUENCIA_ACTIVIDAD","FRECUENCIA_ACTIVIDAD_2","CALIFICACION_DE_PROBABILIDAD_1","CALIFICACION_DEL_IMPACTO_1",
    "CALIFICACION_DE_PROBABILIDAD_2","CALIFICACION_DEL_IMPACTO_2","CALIFICACION_DE_PROBABILIDAD_3","HOMOLOGACION_GRID",
    "CALIFICACION_DE_PROBABILIDAD_4","NIVEL_DE_RIESGO_1","NO_CONTROL","DESCRIPCION_DEL_CONTROL","RESPONSABLE_DEL_CONTROL",
    "AFECTACION","TIPO_DE_CONTROL","IMPLEMENTACION","CALIFICACION","DOCUMENTACION","FRECUENCIA","EVIDENCIA",
    "%_PROBABILIDAD_RESIDUAL","PROBABILIDAD_RESIDUAL_FINAL","%","IMPACTO_RESIDUAL_FINAL","%_IMPACTO_RESIDUAL",
    "ZONA_DE_RIESGO_FINAL","%_CALIFICACION_TOTAL_PROBABILIDAD","PROBABILIDAD_RESIDUAL_TOTAL_FINAL",
    "%_CALIFICACION_TOTAL_IMPACTO","IMPACTO_RESIDUAL_TOTAL_FINAL","CALIFICACION_PROBABILIDAD_RESIDUAL_TOTAL_FINAL",
    "CALIFICACION_IMPACTO_RESIDUAL_TOTAL_FINAL","CALIFICACION_PROBABILIDAD_RESIDUAL","HOMOLOGACION_GRID_2",
    "NIVEL_DE_RIESGO_2","OPCION_DE_TRATAMIENTO","ACCIONES_PARA_MITIGAR_EL_RIESGO_","FECHA_DE_EJECUCION","RESPONSABLE",
    "EFICACIA_DE_LAS_ACCIONES_SINO","SEGUIMIENTO_1","FECHA_1","SEGUIMIENTO_2","FECHA_2","SEGUIMIENTO_3","FECHA_3",
    "SEGUIMIENTO_4","FECHA_4","SEGUIMIENTO_5","FECHA_5","SEGUIMIENTO_6","FECHA_6","SEGUIMIENTO_7"
]

# ============================ Helpers ============================
def _str_clean(s: Optional[str]) -> str:
    if s is None:
        return ""
    s = str(s).replace("\r", " ").replace("\n", " ").replace("\t", " ").strip()
    if PRESERVAR_ENYE:
        s = s.replace("ñ","__enye__").replace("Ñ","__ENYE__")
        s = unidecode(s)
        s = s.replace("__enye__","ñ").replace("__ENYE__","Ñ")
    else:
        s = unidecode(s)  # tildes fuera; ñ->n
    # Mantener '.' por decimales
    s = re.sub(r"[^\x20-\x7E]", "", s)  # quitar no-ASCII visibles
    s = s.translate(str.maketrans({
        '"':"", "'":"", ";":"", "!":"", "¡":"", "*":"", "(": "", ")": ""
    }))  # NO quitamos '.'
    if COLLAPSE_SPACES:
        s = re.sub(r"\s{2,}", " ", s)
    return s.strip()

def _normalize_headers(headers: List[str]) -> List[str]:
    norm, seen = [], set()
    for h in headers:
        h2 = _str_clean(h).lower()
        h2 = re.sub(r"[^a-z0-9_% ]", "", h2).replace(" ", "_")  # conservamos '%' si aparece
        if not h2:
            h2 = "columna_sin_nombre"
        base = h2; k = 1
        while h2 in seen:
            k += 1
            h2 = f"{base}_{k}"
        seen.add(h2)
        norm.append(h2.upper())
    return norm

def read_with_header_row(xlsx_path: Path, sheet: str, header_row_number: int) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet}'. Hojas: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    hdr_idx0 = max(0, header_row_number - 1)
    headers = [str(c) if c is not None else "" for c in rows[hdr_idx0]]
    data = rows[hdr_idx0 + 1:]
    df = pd.DataFrame(data, columns=headers)
    df.dropna(axis=1, how="all", inplace=True)
    df.dropna(axis=0, how="all", inplace=True)
    return df

def autosize_columns(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str):
    ws = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns):
        max_len = max([len(str(col))] + [len(str(x)) for x in df[col].head(200).values])
        ws.set_column(idx, idx, min(max_len + 2, 60))

def find_col(cols: List[str], patrones: List[str]) -> Optional[str]:
    for c in cols:
        for p in patrones:
            if re.match(p, c):
                return c
    return None

def is_percent_col(colname: str) -> bool:
    n = colname.upper()
    if "%" in n:
        return True
    patrones = [
        r".*PORCENTAJE.*",
        r".*%.*",
        r".*PROBABILIDAD_.*RESIDUAL.*",
        r".*IMPACTO_.*RESIDUAL.*",
        r".*CALIFICACION_.*PROBABILIDAD.*",
        r".*CALIFICACION_.*IMPACTO.*",
    ]
    return any(re.match(p, n) for p in patrones)

def parse_percent_value(v: str) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    s = re.sub(r"[^0-9\.,%-]", "", s)
    s = s.replace(" ", "")
    has_pct = "%" in s
    s = s.replace("%", "").replace(",", ".")
    if s.count(".") > 1:
        parts = s.split(".")
        s = "".join(parts[:-1]) + "." + parts[-1]
    try:
        val = float(s)
    except ValueError:
        return ""
    if has_pct or val <= 1.0:
        val = val * (100.0 if val <= 1.0 else 1.0)
    return f"{val:.2f}"

# ============================ Main ============================
def main():
    print(f"[INFO] Leyendo '{INPUT_XLSX.name}' | Hoja '{SHEET_NAME}' | Encabezados fila {HEADER_ROW_NUMBER}")
    df = read_with_header_row(INPUT_XLSX, SHEET_NAME, HEADER_ROW_NUMBER)
    print(f"[INFO] Filas leídas: {len(df)} | Columnas: {len(df.columns)}")

    # Normaliza encabezados
    df.columns = _normalize_headers([str(c) for c in df.columns])

    # Limpia celdas a texto
    for c in df.columns:
        df[c] = df[c].apply(_str_clean)

    # === Identificar columnas clave ===
    col_no   = find_col(df.columns.tolist(), [r"^NO$", r"^NRO$", r"^NUMERO$", r"^ITEM$", r"^CODIGO$"])
    if not col_no:
        raise RuntimeError(f"No se encontró la columna 'No'. Encabezados: {df.columns.tolist()}")

    col_desc = find_col(df.columns.tolist(), [r"^DESCRIPCION_*DEL_*CONTROL$", r"^DESCRIPCION_*CONTROL$", r".*DESCRIPCION.*CONTROL.*"])
    if not col_desc:
        raise RuntimeError(f"No se encontró la columna 'Descripcion del Control'. Encabezados: {df.columns.tolist()}")

    col_subs = find_col(df.columns.tolist(), [r"^SUBSISTEMA_A_LAS_QUE_PERTENECE_EL_RIESGO$"])
    # col_subs puede no existir en algunos archivos; si existe aplicamos la preservación de vacíos

    # Convertir vacíos a NA para poder ffill
    df.replace({"": pd.NA, "None": pd.NA, "nan": pd.NA, "NaN": pd.NA, "NaT": pd.NA, "Nat": pd.NA}, inplace=True)

    # === PRESERVAR VACÍOS ORIGINALES EN SUBSISTEMA... ===
    if col_subs:
        # Máscara de celdas que estaban vacías ANTES del ffill
        mask_subs_vacio = df[col_subs].isna()
    else:
        mask_subs_vacio = None

    # RELLENAR TODAS LAS COLUMNAS (ffill global)
    df = df.ffill(axis=0)

    # Si la columna SUBSISTEMA... existe, restaurar vacíos donde originalmente no había valor
    if col_subs and mask_subs_vacio is not None:
        df.loc[mask_subs_vacio, col_subs] = pd.NA  # restaurar NA (quedará vacío luego)

    # Volver a strings dejando vacíos como ""
    df = df.fillna("")

    # === Filtrar por descripción no vacía ===
    antes = len(df)
    df = df[df[col_desc].astype(str).str.strip() != ""].copy()
    print(f"[INFO] Filtrado '{col_desc}' no vacío: {antes} -> {len(df)} filas")

    # === DEDUPE por (No, Descripción) ===
    df["_DESC_KEY_"] = df[col_desc].str.upper().str.strip()
    df["_NO_KEY_"] = df[col_no].str.upper().str.strip()
    df["_ORD_"] = range(1, len(df) + 1)
    df = df.sort_values("_ORD_").drop_duplicates(subset=["_NO_KEY_", "_DESC_KEY_"], keep="first")
    df.drop(columns=["_DESC_KEY_", "_NO_KEY_", "_ORD_"], inplace=True)
    df = df.reset_index(drop=True)

    # === Normalización de columnas de porcentaje ===
    percent_cols = [c for c in df.columns if is_percent_col(c)]
    for c in percent_cols:
        df[c] = df[c].apply(parse_percent_value)

    # === Renombrado final según listado del usuario ===
    if RENAME_STRICT:
        if len(df.columns) != len(COLUMNAS_FINALES):
            raise RuntimeError(
                f"Número de columnas no coincide para renombrar.\n"
                f"DataFrame: {len(df.columns)} | Esperadas: {len(COLUMNAS_FINALES)}\n"
                f"Encabezados DF: {df.columns.tolist()}"
            )
        df.columns = COLUMNAS_FINALES
    else:
        min_len = min(len(df.columns), len(COLUMNAS_FINALES))
        nuevos = list(COLUMNAS_FINALES[:min_len]) + list(df.columns[min_len:])
        if len(df.columns) < len(COLUMNAS_FINALES):
            for extra in COLUMNAS_FINALES[len(df.columns):]:
                df[extra] = ""
            nuevos = COLUMNAS_FINALES
        df.columns = nuevos

    # === Exportar ===
    with pd.ExcelWriter(OUTPUT_XLSX, engine="xlsxwriter") as writer:
        sheet_out = "MATRIZ_RIESGOS_LIMPIA"
        df.to_excel(writer, index=False, sheet_name=sheet_out)
        autosize_columns(writer, df, sheet_out)

    print(f"[OK] Generado: {OUTPUT_XLSX} | Filas: {len(df)} | Columnas: {len(df.columns)}")
    if percent_cols:
        print(f"[INFO] Columnas tratadas como %: {percent_cols}")
    if col_subs:
        vacios_subs = (df["SUBSISTEMA_A_LAS_QUE_PERTENECE_EL_RIESGO"] == "").sum()
        print(f"[INFO] SUBSISTEMA... vacíos preservados: {vacios_subs}")

if __name__ == "__main__":
    main()